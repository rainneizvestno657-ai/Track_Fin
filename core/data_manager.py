import json
import csv
import os
import sqlite3
import hashlib
from datetime import datetime

from core.locale_manager import LocaleManager

_DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "finance.db")


class DataManager:
    # Exchange rates: how many SOM per 1 unit of foreign currency
    EXCHANGE_RATES: dict[str, float] = {
        "usd": 87.47,
        "rub": 1.24,
    }
    CURRENCY_SYMBOLS: dict[str, str] = {
        "som": "с",
        "usd": "$",
        "rub": "₽",
    }

    def __init__(self):
        # Locale / i18n
        self.locale = LocaleManager(default="ru")
        # Active display currency ("som" | "usd" | "rub")
        self.currency: str = "som"
        # Active stats period shown on Transactions screen ("day"|"week"|"month"|"year")
        self.active_period: str = "month"

        self.monthly_income   = 0.0
        self.monthly_expenses = 0.0
        self.balance          = 0.0

        # Single goal card shown on transactions dashboard
        self.goal_name  = ""
        self.goal_total = 0.0
        self.goal_saved = 0.0
        # Index of the goal pinned to the Transactions card (-1 = none)
        self.active_goal_index: int = -1

        # Category expenses used on transactions dashboard
        self.category_expenses: list[dict] = []

        # Full categories with spending limits (Categories screen)
        self.categories = [
            {"key": "cat_default_food",          "name": self.locale.t("cat_default_food"),          "icon": "🍔", "limit": 0.0, "spent": 0.0},
            {"key": "cat_default_transport",     "name": self.locale.t("cat_default_transport"),     "icon": "🚗", "limit": 0.0, "spent": 0.0},
            {"key": "cat_default_entertainment", "name": self.locale.t("cat_default_entertainment"), "icon": "🎮", "limit": 0.0, "spent": 0.0},
            {"key": "cat_default_subscriptions", "name": self.locale.t("cat_default_subscriptions"), "icon": "💳", "limit": 0.0, "spent": 0.0},
            {"key": "cat_default_mobile",        "name": self.locale.t("cat_default_mobile"),        "icon": "📱", "limit": 0.0, "spent": 0.0},
            {"key": "cat_default_other",         "name": self.locale.t("cat_default_other"),         "icon": "📦", "limit": 0.0, "spent": 0.0},
        ]

        # Goals (Goals screen)
        self.goals: list[dict] = []

        # Savings history shown on Goals screen
        self.savings_history: list[dict] = []

        # Analytics: previous month totals
        self.prev_monthly_income   = 0.0
        self.prev_monthly_expenses = 0.0

        # Analytics: per-category comparison (previous month vs this month)
        self.analytics_comparison: list[dict] = []

        # Transaction history
        self.transactions: list[dict] = []

        # Saved transfer drafts (phone + recipient_id for visual reference)
        self.drafts: list[dict] = []

        # ── Database ─────────────────────────────────────────────
        os.makedirs(os.path.dirname(_DB_PATH), exist_ok=True)
        self._db_path = _DB_PATH
        self._init_db()
        self._load()

    # ──────────────────────────────────────────
    # Database — init
    # ──────────────────────────────────────────

    def _init_db(self):
        con = sqlite3.connect(self._db_path)
        con.executescript("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS transactions (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                date       TEXT NOT NULL,
                vendor     TEXT NOT NULL,
                category   TEXT NOT NULL,
                comment    TEXT NOT NULL,
                amount     REAL NOT NULL,
                type       TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS categories (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                sort_order   INTEGER NOT NULL,
                cat_key      TEXT,
                name         TEXT NOT NULL,
                icon         TEXT NOT NULL,
                limit_amount REAL NOT NULL,
                spent        REAL NOT NULL
            );
            CREATE TABLE IF NOT EXISTS category_expenses (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                amount   REAL NOT NULL
            );
            CREATE TABLE IF NOT EXISTS goals (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                sort_order  INTEGER NOT NULL,
                name        TEXT NOT NULL,
                icon        TEXT NOT NULL,
                target      REAL NOT NULL,
                saved       REAL NOT NULL,
                months_left INTEGER NOT NULL
            );
            CREATE TABLE IF NOT EXISTS savings_history (
                id     INTEGER PRIMARY KEY AUTOINCREMENT,
                date   TEXT NOT NULL,
                amount REAL NOT NULL,
                goal   TEXT NOT NULL,
                type   TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS analytics_comparison (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                prev     REAL NOT NULL,
                curr     REAL NOT NULL
            );
            CREATE TABLE IF NOT EXISTS drafts (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                phone        TEXT NOT NULL,
                recipient_id TEXT NOT NULL,
                amount       REAL NOT NULL,
                date         TEXT NOT NULL
            );
        """)
        # Migration: add created_at to existing transactions table
        try:
            con.execute("ALTER TABLE transactions ADD COLUMN created_at TEXT NOT NULL DEFAULT ''")
            con.commit()
        except sqlite3.OperationalError:
            pass
        con.commit()
        con.close()

    # ──────────────────────────────────────────
    # Database — load
    # ──────────────────────────────────────────

    def _load(self):
        con = sqlite3.connect(self._db_path)
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        # Settings (scalars)
        cur.execute("SELECT key, value FROM settings")
        s = {row["key"]: row["value"] for row in cur.fetchall()}
        if s:
            self.balance               = float(s.get("balance", 0.0))
            self.monthly_income        = float(s.get("monthly_income", 0.0))
            self.monthly_expenses      = float(s.get("monthly_expenses", 0.0))
            self.prev_monthly_income   = float(s.get("prev_monthly_income", 0.0))
            self.prev_monthly_expenses = float(s.get("prev_monthly_expenses", 0.0))
            self.currency              = s.get("currency", "som")
            self.locale.set_lang(s.get("language", "ru"))
            self.active_period         = s.get("active_period", "month")
            self.active_goal_index     = int(s.get("active_goal_index", -1))

        # Transactions (newest first — stored newest first)
        cur.execute("SELECT date, vendor, category, comment, amount, type, COALESCE(created_at,'') as created_at FROM transactions ORDER BY id DESC")
        self.transactions = [
            {"date": r["date"], "vendor": r["vendor"], "category": r["category"],
             "comment": r["comment"], "amount": r["amount"], "type": r["type"],
             "created_at": r["created_at"]}
            for r in cur.fetchall()
        ]

        # Categories
        cur.execute("SELECT cat_key, name, icon, limit_amount, spent FROM categories ORDER BY sort_order")
        rows = cur.fetchall()
        if rows:
            self.categories = []
            for r in rows:
                cat = {
                    "name":  self.locale.t(r["cat_key"]) if r["cat_key"] else r["name"],
                    "icon":  r["icon"],
                    "limit": r["limit_amount"],
                    "spent": r["spent"],
                }
                if r["cat_key"]:
                    cat["key"] = r["cat_key"]
                self.categories.append(cat)

        # Category expenses (dashboard breakdown)
        cur.execute("SELECT category, amount FROM category_expenses ORDER BY id")
        self.category_expenses = [
            {"category": r["category"], "amount": r["amount"]}
            for r in cur.fetchall()
        ]

        # Goals
        cur.execute("SELECT name, icon, target, saved, months_left FROM goals ORDER BY sort_order")
        self.goals = [
            {"name": r["name"], "icon": r["icon"], "target": r["target"],
             "saved": r["saved"], "months_left": r["months_left"]}
            for r in cur.fetchall()
        ]
        # Sync pinned goal fields after loading
        self._sync_active_goal()

        # Savings history (newest first)
        cur.execute("SELECT date, amount, goal, type FROM savings_history ORDER BY id DESC")
        self.savings_history = [
            {"date": r["date"], "amount": r["amount"], "goal": r["goal"], "type": r["type"]}
            for r in cur.fetchall()
        ]

        # Analytics comparison
        cur.execute("SELECT category, prev, curr FROM analytics_comparison ORDER BY id")
        self.analytics_comparison = [
            {"category": r["category"], "prev": r["prev"], "curr": r["curr"]}
            for r in cur.fetchall()
        ]

        # Transfer drafts (newest first)
        try:
            cur.execute("SELECT phone, recipient_id, amount, date FROM drafts ORDER BY id DESC")
            self.drafts = [
                {"phone": r["phone"], "recipient_id": r["recipient_id"],
                 "amount": r["amount"], "date": r["date"]}
                for r in cur.fetchall()
            ]
        except Exception:
            self.drafts = []

        con.close()

        # ── Auto-clean stale category_expenses ────────────────────────────────
        # If a category was deleted before the cleanup fix existed, its entry
        # may still be persisted in the DB.  Remove any entries whose category
        # name no longer matches any current category.
        active_names = {c["name"].lower() for c in self.categories}
        clean = [ce for ce in self.category_expenses
                 if ce["category"].lower() in active_names]
        if len(clean) < len(self.category_expenses):
            self.category_expenses = clean
            self._save_category_expenses()

    # ──────────────────────────────────────────
    # Database — save helpers
    # ──────────────────────────────────────────

    def _save_settings(self):
        con = sqlite3.connect(self._db_path)
        con.executemany(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            [
                ("balance",               str(self.balance)),
                ("monthly_income",        str(self.monthly_income)),
                ("monthly_expenses",      str(self.monthly_expenses)),
                ("prev_monthly_income",   str(self.prev_monthly_income)),
                ("prev_monthly_expenses", str(self.prev_monthly_expenses)),
                ("currency",              self.currency),
                ("language",              self.locale.current),
                ("active_period",         self.active_period),
                ("active_goal_index",     str(self.active_goal_index)),
            ],
        )
        con.commit()
        con.close()

    def _save_transactions(self):
        con = sqlite3.connect(self._db_path)
        con.execute("DELETE FROM transactions")
        con.executemany(
            "INSERT INTO transactions (date, vendor, category, comment, amount, type, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            [(t["date"], t["vendor"], t["category"], t["comment"], t["amount"], t["type"], t.get("created_at", ""))
             for t in self.transactions],
        )
        con.commit()
        con.close()

    def _save_drafts(self):
        con = sqlite3.connect(self._db_path)
        con.execute("DELETE FROM drafts")
        con.executemany(
            "INSERT INTO drafts (phone, recipient_id, amount, date) VALUES (?, ?, ?, ?)",
            [(d["phone"], d["recipient_id"], d["amount"], d["date"]) for d in self.drafts],
        )
        con.commit()
        con.close()

    def _save_categories(self):
        con = sqlite3.connect(self._db_path)
        con.execute("DELETE FROM categories")
        con.executemany(
            "INSERT INTO categories (sort_order, cat_key, name, icon, limit_amount, spent) VALUES (?, ?, ?, ?, ?, ?)",
            [(i, cat.get("key"), cat["name"], cat["icon"], cat["limit"], cat["spent"])
             for i, cat in enumerate(self.categories)],
        )
        con.commit()
        con.close()

    def _save_category_expenses(self):
        con = sqlite3.connect(self._db_path)
        con.execute("DELETE FROM category_expenses")
        con.executemany(
            "INSERT INTO category_expenses (category, amount) VALUES (?, ?)",
            [(c["category"], c["amount"]) for c in self.category_expenses],
        )
        con.commit()
        con.close()

    def _save_goals(self):
        con = sqlite3.connect(self._db_path)
        con.execute("DELETE FROM goals")
        con.executemany(
            "INSERT INTO goals (sort_order, name, icon, target, saved, months_left) VALUES (?, ?, ?, ?, ?, ?)",
            [(i, g["name"], g["icon"], g["target"], g["saved"], g["months_left"])
             for i, g in enumerate(self.goals)],
        )
        con.commit()
        con.close()

    def _save_savings_history(self):
        con = sqlite3.connect(self._db_path)
        con.execute("DELETE FROM savings_history")
        con.executemany(
            "INSERT INTO savings_history (date, amount, goal, type) VALUES (?, ?, ?, ?)",
            [(s["date"], s["amount"], s["goal"], s["type"]) for s in self.savings_history],
        )
        con.commit()
        con.close()

    def _save_analytics(self):
        con = sqlite3.connect(self._db_path)
        con.execute("DELETE FROM analytics_comparison")
        con.executemany(
            "INSERT INTO analytics_comparison (category, prev, curr) VALUES (?, ?, ?)",
            [(a["category"], a["prev"], a["curr"]) for a in self.analytics_comparison],
        )
        con.commit()
        con.close()

    # ──────────────────────────────────────────
    # Locale helpers
    # ──────────────────────────────────────────

    def set_currency(self, code: str) -> None:
        if code in self.CURRENCY_SYMBOLS:
            self.currency = code
            self._save_settings()

    def set_language(self, code: str) -> None:
        self.locale.set_lang(code)
        self._save_settings()

    def set_period(self, period: str) -> None:
        if period in ("day", "week", "month", "year"):
            self.active_period = period
            self._save_settings()

    def format_amount(self, value: float) -> str:
        """Format *value* (always stored in som) using the active display currency."""
        if self.currency == "som":
            from core.utils import format_kgs
            return format_kgs(value)
        rate = self.EXCHANGE_RATES[self.currency]
        sym  = self.CURRENCY_SYMBOLS[self.currency]
        converted = value / rate
        formatted = f"{converted:,.2f}".replace(",", ".")
        return f"{formatted} {sym}"

    def refresh_category_names(self):
        """Update display names of default categories after a language switch."""
        for cat in self.categories:
            if "key" in cat:
                cat["name"] = self.locale.t(cat["key"])

    def _today(self) -> str:
        return datetime.now().strftime("%d.%m.%Y")

    def _other_category_name(self) -> str:
        for cat in self.categories:
            if cat.get("key") == "cat_default_other":
                return cat["name"]
        return self.locale.t("cat_default_other")

    # ──────────────────────────────────────────
    # Computed helpers
    # ──────────────────────────────────────────

    def _recalculate_balance(self):
        self.balance = self.monthly_income - self.monthly_expenses

    def get_largest_expense_category(self) -> str:
        if not self.category_expenses:
            return ""
        return max(self.category_expenses, key=lambda c: c["amount"])["category"]

    def get_expense_change_percent(self) -> float:
        if self.prev_monthly_expenses == 0:
            return 0.0
        return (self.monthly_expenses - self.prev_monthly_expenses) / self.prev_monthly_expenses * 100

    def get_income_change_percent(self) -> float:
        if self.prev_monthly_income == 0:
            return 0.0
        return (self.monthly_income - self.prev_monthly_income) / self.prev_monthly_income * 100

    def get_largest_category_change_percent(self) -> tuple:
        largest = self.get_largest_expense_category()
        for comp in self.analytics_comparison:
            if comp["category"] == largest and comp["prev"] > 0:
                pct = (comp["curr"] - comp["prev"]) / comp["prev"] * 100
                return largest, pct
        return largest, 0.0

    # ──────────────────────────────────────────
    # Transactions
    # ──────────────────────────────────────────

    def add_deposit(self, amount: float):
        if amount <= 0:
            return
        self.monthly_income += amount
        self._recalculate_balance()
        self.transactions.insert(0, {
            "date": self._today(), "vendor": "Банкомат",
            "category": "Пополнение", "comment": "Быстрое пополнение",
            "amount": amount, "type": "income",
            "created_at": datetime.now().isoformat(),
        })
        self._save_settings()
        self._save_transactions()

    def add_transfer(self, amount: float):
        if amount <= 0:
            return
        self.monthly_expenses += amount
        self._recalculate_balance()
        other = self._other_category_name()
        found = False
        for cat in self.category_expenses:
            if cat["category"] == other:
                cat["amount"] += amount
                found = True
                break
        if not found:
            self.category_expenses.append({"category": other, "amount": amount})
        for cat in self.categories:
            if cat.get("key") == "cat_default_other":
                cat["spent"] += amount
                break
        self.transactions.insert(0, {
            "date": self._today(), "vendor": "Перевод",
            "category": other, "comment": "Быстрый перевод",
            "amount": amount, "type": "expense",
            "created_at": datetime.now().isoformat(),
        })
        self._save_settings()
        self._save_transactions()
        self._save_category_expenses()
        self._save_categories()

    def add_custom_transaction(self, amount: float, tx_type: str, category: str, vendor: str, comment: str):
        if amount <= 0:
            return
        if tx_type == "income":
            self.monthly_income += amount
        else:
            self.monthly_expenses += amount
            found = False
            for cat in self.category_expenses:
                if cat["category"].lower() == category.lower():
                    cat["amount"] += amount
                    found = True
                    break
            if not found:
                self.category_expenses.append({"category": category, "amount": amount})
            for cat in self.categories:
                if cat["name"].lower() == category.lower():
                    cat["spent"] += amount
                    break
        self._recalculate_balance()
        self.transactions.insert(0, {
            "date": self._today(), "vendor": vendor,
            "category": category, "comment": comment,
            "amount": amount, "type": tx_type,
            "created_at": datetime.now().isoformat(),
        })
        self._save_settings()
        self._save_transactions()
        self._save_category_expenses()
        self._save_categories()

    # ──────────────────────────────────────────
    # Categories
    # ──────────────────────────────────────────

    def add_category(self, name: str, icon: str, limit: float):
        self.categories.append({"name": name, "icon": icon, "limit": limit, "spent": 0.0})
        self._save_categories()

    def edit_category(self, index: int, name: str, icon: str, limit: float):
        if 0 <= index < len(self.categories):
            self.categories[index].update({"name": name, "icon": icon, "limit": limit})
            self._save_categories()

    def delete_category(self, index: int):
        if 0 <= index < len(self.categories):
            cat_name = self.categories[index]["name"]
            self.categories.pop(index)
            # Remove related expenses from dashboard breakdown
            self.category_expenses = [
                c for c in self.category_expenses
                if c["category"].lower() != cat_name.lower()
            ]
            self._save_categories()
            self._save_category_expenses()

    # ──────────────────────────────────────────
    # Goals
    # ──────────────────────────────────────────

    def add_goal(self, name: str, icon: str, target: float, months_left: int):
        self.goals.append({"name": name, "icon": icon, "target": target, "saved": 0.0, "months_left": months_left})
        self._save_goals()

    def add_saving_to_goal(self, goal_index: int, amount: float):
        if 0 <= goal_index < len(self.goals) and amount > 0:
            self.goals[goal_index]["saved"] += amount
            self.savings_history.insert(0, {
                "date": self._today(),
                "amount": amount,
                "goal": self.goals[goal_index]["name"],
                "type": "Перевод",
            })
            if goal_index == self.active_goal_index:
                self.goal_saved = self.goals[goal_index]["saved"]
            self._save_goals()
            self._save_savings_history()

    def edit_goal(self, index: int, name: str, icon: str, target: float, months: int):
        if 0 <= index < len(self.goals):
            self.goals[index].update({"name": name, "icon": icon, "target": target, "months_left": months})
            if index == self.active_goal_index:
                self.goal_name  = name
                self.goal_total = target
            self._save_goals()

    def delete_goal(self, index: int):
        if 0 <= index < len(self.goals):
            self.goals.pop(index)
            # Adjust active_goal_index after deletion
            if self.active_goal_index == index:
                self.active_goal_index = 0 if self.goals else -1
            elif self.active_goal_index > index:
                self.active_goal_index -= 1
            self._sync_active_goal()
            self._save_goals()
            self._save_settings()

    def _sync_active_goal(self):
        """Update goal_name/goal_total/goal_saved from active_goal_index."""
        if 0 <= self.active_goal_index < len(self.goals):
            g = self.goals[self.active_goal_index]
            self.goal_name  = g["name"]
            self.goal_total = g["target"]
            self.goal_saved = g["saved"]
        else:
            self.goal_name  = ""
            self.goal_total = 0.0
            self.goal_saved = 0.0

    def set_active_goal(self, index: int):
        """Pin a goal to the Transactions dashboard card and persist."""
        if 0 <= index < len(self.goals):
            self.active_goal_index = index
        else:
            self.active_goal_index = -1
        self._sync_active_goal()
        self._save_settings()

    # ──────────────────────────────────────────
    # Cancel / period stats / drafts
    # ──────────────────────────────────────────

    def cancel_transaction(self, created_at: str) -> bool:
        """Reverse and remove a transaction identified by its created_at ISO string."""
        for i, tx in enumerate(self.transactions):
            if tx.get("created_at") == created_at:
                amount = tx["amount"]
                if tx["type"] == "income":
                    self.monthly_income = max(0.0, self.monthly_income - amount)
                else:
                    self.monthly_expenses = max(0.0, self.monthly_expenses - amount)
                    for cat_exp in self.category_expenses[:]:
                        if cat_exp["category"].lower() == tx["category"].lower():
                            cat_exp["amount"] -= amount
                            if cat_exp["amount"] <= 0:
                                self.category_expenses.remove(cat_exp)
                            break
                    for cat in self.categories:
                        if cat["name"].lower() == tx["category"].lower():
                            cat["spent"] = max(0.0, cat["spent"] - amount)
                            break
                self._recalculate_balance()
                self.transactions.pop(i)
                self._save_settings()
                self._save_transactions()
                self._save_category_expenses()
                self._save_categories()
                return True
        return False

    def get_period_stats(self, period: str) -> dict:
        """Return income/expense totals and percentages for day/week/month/year."""
        from datetime import timedelta, date as date_type
        today = datetime.now().date()

        if period == "day":
            start = today                          # only today
        elif period == "week":
            start = today - timedelta(days=6)      # last 7 days inclusive
        elif period == "month":
            start = today.replace(day=1)           # 1st of current month
        else:                                      # year
            start = today.replace(month=1, day=1)  # 1 Jan of current year

        income = 0.0
        expenses = 0.0
        for tx in self.transactions:
            try:
                tx_date = datetime.strptime(tx["date"], "%d.%m.%Y").date()
            except ValueError:
                continue
            if start <= tx_date <= today:
                if tx["type"] == "income":
                    income += tx["amount"]
                else:
                    expenses += tx["amount"]

        total = income + expenses
        return {
            "income":      income,
            "expenses":    expenses,
            "income_pct":  (income   / total * 100) if total > 0 else 0.0,
            "expense_pct": (expenses / total * 100) if total > 0 else 0.0,
        }

    def _period_range(self, period: str):
        """Return (start, end) date tuple for the current period."""
        from datetime import timedelta
        today = datetime.now().date()
        if period == "day":
            return today, today
        elif period == "week":
            return today - timedelta(days=6), today
        elif period == "month":
            return today.replace(day=1), today
        else:
            return today.replace(month=1, day=1), today

    def _prev_period_range(self, period: str):
        """Return (start, end) date tuple for the PREVIOUS equivalent period."""
        from datetime import timedelta
        today = datetime.now().date()
        if period == "day":
            d = today - timedelta(days=1)
            return d, d
        elif period == "week":
            end   = today - timedelta(days=7)
            return end - timedelta(days=6), end
        elif period == "month":
            end   = today.replace(day=1) - timedelta(days=1)
            return end.replace(day=1), end
        else:
            prev = today.year - 1
            from datetime import date
            return date(prev, 1, 1), date(prev, 12, 31)

    def get_prev_period_stats(self, period: str) -> dict:
        """Income/expense totals for the PREVIOUS equivalent period."""
        start, end = self._prev_period_range(period)
        income = expenses = 0.0
        for tx in self.transactions:
            try:
                d = datetime.strptime(tx["date"], "%d.%m.%Y").date()
            except ValueError:
                continue
            if start <= d <= end:
                if tx["type"] == "income":
                    income += tx["amount"]
                else:
                    expenses += tx["amount"]
        total = income + expenses
        return {
            "income":      income,
            "expenses":    expenses,
            "income_pct":  (income   / total * 100) if total > 0 else 0.0,
            "expense_pct": (expenses / total * 100) if total > 0 else 0.0,
        }

    def get_period_category_expenses(self, period: str) -> list:
        """Per-category expense list for the current period, sorted by amount desc."""
        start, end = self._period_range(period)
        totals: dict = {}
        for tx in self.transactions:
            if tx["type"] != "expense":
                continue
            try:
                d = datetime.strptime(tx["date"], "%d.%m.%Y").date()
            except ValueError:
                continue
            if start <= d <= end:
                totals[tx["category"]] = totals.get(tx["category"], 0.0) + tx["amount"]
        return sorted(
            [{"category": c, "amount": a} for c, a in totals.items()],
            key=lambda x: -x["amount"],
        )

    def get_period_comparison(self, period: str) -> list:
        """Per-category {prev, curr} comparison for analytics table."""
        curr_start, curr_end = self._period_range(period)
        prev_start, prev_end = self._prev_period_range(period)
        curr: dict = {}
        prev: dict = {}
        for tx in self.transactions:
            if tx["type"] != "expense":
                continue
            try:
                d = datetime.strptime(tx["date"], "%d.%m.%Y").date()
            except ValueError:
                continue
            if curr_start <= d <= curr_end:
                curr[tx["category"]] = curr.get(tx["category"], 0.0) + tx["amount"]
            elif prev_start <= d <= prev_end:
                prev[tx["category"]] = prev.get(tx["category"], 0.0) + tx["amount"]
        all_cats = set(curr) | set(prev)
        rows = [{"category": c, "prev": prev.get(c, 0.0), "curr": curr.get(c, 0.0)}
                for c in all_cats if curr.get(c, 0.0) > 0 or prev.get(c, 0.0) > 0]
        return sorted(rows, key=lambda x: -x["curr"])

    def save_draft(self, phone: str, recipient_id: str, amount: float):
        self.drafts.insert(0, {
            "phone": phone, "recipient_id": recipient_id,
            "amount": amount, "date": self._today(),
        })
        self._save_drafts()

    def delete_draft(self, index: int):
        if 0 <= index < len(self.drafts):
            self.drafts.pop(index)
            self._save_drafts()

    # ──────────────────────────────────────────
    # Export & Encryption
    # ──────────────────────────────────────────

    def _encrypt_file(self, path: str, password: str) -> str:
        """Encrypt *path* with *password* and return the encrypted file path.
        For .xlsx files uses Excel-native password protection via msoffcrypto-tool
        so Excel will ask for the password on open.  Falls back to XOR for other
        formats or when msoffcrypto-tool is not installed."""
        if path.endswith(".xlsx"):
            try:
                import msoffcrypto
                tmp_path = path + ".tmp"
                with open(path, "rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    with open(tmp_path, "wb") as g:
                        office_file.encrypt(password, g)
                os.remove(path)
                os.rename(tmp_path, path)
                return path
            except ImportError:
                pass
        # Fallback: XOR encryption for JSON/CSV or when msoffcrypto is missing
        key = hashlib.sha256(password.encode("utf-8")).digest()
        with open(path, "rb") as f:
            data = f.read()
        encrypted = bytes(data[i] ^ key[i % len(key)] for i in range(len(data)))
        enc_path = path + ".enc"
        with open(enc_path, "wb") as f:
            f.write(encrypted)
        os.remove(path)
        return enc_path

    def export_data(self, what: list, fmt: str, folder: str, password: str = "") -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        data = {}
        if "transactions" in what:
            data["transactions"] = self.transactions
        if "categories" in what:
            data["categories"] = self.categories
        if "goals" in what:
            data["goals"] = self.goals
        if "balance" in what:
            data["balance"] = {
                "current": self.balance,
                "monthly_income": self.monthly_income,
                "monthly_expenses": self.monthly_expenses,
            }

        if fmt == "json":
            path = os.path.join(folder, f"finance_export_{timestamp}.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        elif fmt == "csv":
            path = os.path.join(folder, f"finance_export_{timestamp}.csv")
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                if "transactions" in data:
                    w.writerow(["=== ТРАНЗАКЦИИ ==="])
                    w.writerow(["Дата", "Контрагент", "Категория", "Комментарий", "Сумма", "Тип"])
                    for tx in data["transactions"]:
                        w.writerow([tx["date"], tx["vendor"], tx["category"],
                                    tx["comment"], tx["amount"], tx["type"]])
                if "categories" in data:
                    w.writerow([])
                    w.writerow(["=== КАТЕГОРИИ ==="])
                    w.writerow(["Название", "Иконка", "Лимит", "Потрачено"])
                    for cat in data["categories"]:
                        w.writerow([cat["name"], cat["icon"], cat["limit"], cat["spent"]])
                if "goals" in data:
                    w.writerow([])
                    w.writerow(["=== ЦЕЛИ ==="])
                    w.writerow(["Название", "Цель (с)", "Накоплено (с)", "Осталось месяцев"])
                    for g in data["goals"]:
                        w.writerow([g["name"], g["target"], g["saved"], g["months_left"]])
                if "balance" in data:
                    w.writerow([])
                    w.writerow(["=== БАЛАНС ==="])
                    b = data["balance"]
                    w.writerow(["Текущий баланс", "Доходы за месяц", "Расходы за месяц"])
                    w.writerow([b["current"], b["monthly_income"], b["monthly_expenses"]])

        elif fmt == "excel":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                path = os.path.join(folder, f"finance_export_{timestamp}.xlsx")
                wb = openpyxl.Workbook()
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="1E351F")

                def _write_sheet(ws, headers, rows):
                    ws.append(headers)
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    for row in rows:
                        ws.append(row)
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 20

                if "transactions" in data:
                    ws = wb.active
                    ws.title = "Транзакции"
                    _write_sheet(ws,
                        ["Дата", "Контрагент", "Категория", "Комментарий", "Сумма", "Тип"],
                        [[t["date"], t["vendor"], t["category"], t["comment"], t["amount"], t["type"]]
                         for t in data["transactions"]])
                if "categories" in data:
                    ws2 = wb.create_sheet("Категории")
                    _write_sheet(ws2,
                        ["Название", "Иконка", "Лимит (с)", "Потрачено (с)"],
                        [[c["name"], c["icon"], c["limit"], c["spent"]] for c in data["categories"]])
                if "goals" in data:
                    ws3 = wb.create_sheet("Цели")
                    _write_sheet(ws3,
                        ["Название", "Цель (с)", "Накоплено (с)", "Осталось месяцев"],
                        [[g["name"], g["target"], g["saved"], g["months_left"]] for g in data["goals"]])
                if "balance" in data:
                    ws4 = wb.create_sheet("Баланс")
                    b = data["balance"]
                    _write_sheet(ws4,
                        ["Текущий баланс (с)", "Доходы за месяц (с)", "Расходы за месяц (с)"],
                        [[b["current"], b["monthly_income"], b["monthly_expenses"]]])
                wb.save(path)
            except ImportError:
                path = os.path.join(folder, f"finance_export_{timestamp}_excel.csv")
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f)
                    w.writerow(["openpyxl не установлен. Экспортировано в CSV формате."])

        if password:
            path = self._encrypt_file(path, password)

        return path
